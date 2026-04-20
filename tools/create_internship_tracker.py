from pathlib import Path
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


def build_workbook() -> Workbook:
    headers = [
        "投递公司",
        "公司类型",
        "投递岗位",
        "投递链接",
        "投递日期",
        "工作地点",
        "是否已测评",
        "是否已笔试",
        "是否已面试",
        "是否有offer",
        "其它",
        "总结",
    ]

    sample_rows = [
        [
            "拼多多",
            "互联网",
            "Java后端开发实习生",
            "https://example.com",
            "",
            "上海",
            "否",
            "否",
            "否",
            "否",
            "内推/官网",
            "优先关注后端与中间件方向",
        ],
        [
            "同花顺",
            "金融科技",
            "Java开发实习生",
            "https://example.com",
            "",
            "杭州",
            "否",
            "否",
            "否",
            "否",
            "可关注暑期转正机会",
            "结合你的后端项目匹配度较高",
        ],
        [
            "科大讯飞",
            "人工智能",
            "大模型应用开发实习生（Java）",
            "https://example.com",
            "",
            "合肥/北京",
            "否",
            "否",
            "否",
            "否",
            "关注 Java + AI 方向",
            "与你的智能客服与模型部署经历较贴合",
        ],
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.freeze_panes = "A3"

    ws["A1"] = "暑期实习投递汇总"
    ws["A1"].font = Font(size=15, bold=True, color="1F1F1F")
    ws.merge_cells("A1:L1")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    ws["A2"] = "说明：填写投递进度后，可直接按列筛选；状态列建议统一使用 是/否/待定。"
    ws["A2"].font = Font(size=10, color="9C3D00")
    ws["A2"].fill = PatternFill("solid", fgColor="FCE4D6")
    ws.merge_cells("A2:L2")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 28

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="1F1F1F")
    thin = Side(style="thin", color="D0D7DE")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(sample_rows, start=4):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = cell_border
            if col_idx == 4 and value:
                cell.style = "Hyperlink"
                cell.hyperlink = value
            if col_idx in {7, 8, 9, 10}:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

    status_dv = DataValidation(type="list", formula1='"是,否,待定"', allow_blank=True)
    ws.add_data_validation(status_dv)

    for row_idx in range(4, 205):
        for col_idx in range(1, 13):
            ws.cell(row=row_idx, column=col_idx).border = cell_border
        for col_idx in (7, 8, 9, 10):
            status_dv.add(ws.cell(row=row_idx, column=col_idx))
        ws.cell(row=row_idx, column=5).number_format = "yyyy-mm-dd"

    col_widths = {
        "A": 18,
        "B": 14,
        "C": 24,
        "D": 34,
        "E": 13,
        "F": 14,
        "G": 12,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 24,
        "L": 30,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    for row in range(3, 205):
        ws.row_dimensions[row].height = 22

    table = Table(displayName="InternApplyTable", ref="A3:L204")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    help_ws = wb.create_sheet("Help")
    help_ws["A1"] = "字段说明"
    help_ws["A1"].font = Font(size=14, bold=True)

    help_rows = [
        ("投递公司", "公司名称"),
        ("公司类型", "如 互联网/金融科技/人工智能/制造业"),
        ("投递岗位", "具体岗位名称，建议保留 JD 原文"),
        ("投递链接", "官网/内推/招聘平台链接"),
        ("投递日期", "实际投递日期"),
        ("工作地点", "城市或远程"),
        ("是否已测评", "填写 是/否/待定"),
        ("是否已笔试", "填写 是/否/待定"),
        ("是否已面试", "填写 是/否/待定"),
        ("是否有offer", "填写 是/否/待定"),
        ("其它", "内推码、备注、联系人、截止时间等"),
        ("总结", "记录匹配度、准备重点、下一步动作"),
    ]
    for idx, (key, value) in enumerate(help_rows, start=3):
        help_ws[f"A{idx}"] = key
        help_ws[f"B{idx}"] = value
        help_ws[f"A{idx}"].font = Font(bold=True)

    help_ws.column_dimensions["A"].width = 18
    help_ws.column_dimensions["B"].width = 48

    return wb


def main() -> None:
    desktop = Path(os.path.expanduser("~")) / "Desktop"
    out_path = desktop / "暑期实习投递汇总.xlsx"
    workbook = build_workbook()
    workbook.save(out_path)
    print(out_path)


if __name__ == "__main__":
    main()
